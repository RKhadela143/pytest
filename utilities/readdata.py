import openpyxl

 
# Description: To read username password from the excel file
def readdata(sheetname, bookname):
    list1 = []
    book = openpyxl.load_workbook(bookname)
    sheet = book[sheetname]
    rows = sheet.max_row
    # cols = sheet.max_column

    for r in range(2, rows+1):
        username = sheet.cell(r, 1).value
        password = sheet.cell(r, 2).value

        tuple1 = (username, password)
        list1.append(tuple1)
    print(list1)
    return list1

# Description: To read start date and end date
def read_date(sheetname, bookname):
    mydict = {}
    book = openpyxl.load_workbook(bookname)
    sheet = book[sheetname]
    rows = sheet.max_row
    cols = sheet.max_column
    mydict["startdate"] = sheet.cell(2, 1).value
    mydict["enddate"] = sheet.cell(2, 2).value

    return mydict

# Description: To read browsers details to execute tests
def readbrowser(sheetname, bookname):
    list2 = []
    book = openpyxl.load_workbook(bookname)
    sheet = book[sheetname]
    rows = sheet.max_row
    # cols = sheet.max_column

    for r in range(2, rows+1):
        browser = sheet.cell(r, 1).value
        list2.append(browser)
    print(list2)
    return list2

